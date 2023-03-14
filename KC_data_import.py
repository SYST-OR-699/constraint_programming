
import pandas as pd

from openpyxl import load_workbook

# enter the data

sheet_names = pd.ExcelFile("small_inputs_gmuV2_msf2.xlsx").sheet_names
trial_names = [sheet_names[x] for x in [3, 6, 18]]


# using openpyxl
# wb = load_workbook(filename="small_inputs_gmuV2_msf2.xlsx")

# wb.sheetnames

# d = pd.read_excel("trial.xlsx", sheet_name=sheet_names[0],
#                   usecols="A:H", nrows=30)

d = pd.read_excel("small_inputs_gmuV2_msf2.xlsx", sheet_name=trial_names, keep_default_na=False)

# d = []
# # filename = "trial.xlsx"
# filename = "small_inputs_gmuV2_msf2.xlsx"

# for sheet in trial_names:
#     for row in wb[sheet].iter_rows(min_row=1, min_col=1, values_only=True):
#         d.append(pd.read_excel(filename, sheet_name=sheet))

# define indices


# not sure best way to do this yet. Do we distill the excel file to a csv?



# define variables

# define objective function????

# define constraints